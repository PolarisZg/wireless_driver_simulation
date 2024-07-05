import openpyxl

x86_config_path = "/home/yukikaze/code/wireless_driver_simulation/Tools/x86_64/config-6.5.0-35-generic"
riscv_config_path = "/home/yukikaze/code/wireless_driver_simulation/Tools/riscv/config-6.8.0-35-generic"

config_dict = {
    'key' : ['value1', 'value1'],
}

with open(x86_config_path, "r", encoding="utf-8") as x86_config_file :
    for line in x86_config_file :
        line = line.strip()
        if '=' in line : 
            if(len(line) > 2) :
                if(line[0] == '#') :
                    line = line[1:]
                parts = line.split('=')
                config_dict[parts[0]] = ['nouse', 'nouse']

with open(riscv_config_path, "r", encoding="utf-8") as riscv_config_file :
    for line in riscv_config_file :
        line = line.strip()
        if '=' in line : 
            if(len(line) > 2) :
                if(line[0] == '#') :
                    line = line[1:]
                parts = line.split('=')
                config_dict[parts[0]] = ['nouse', 'nouse']

with open(x86_config_path, "r", encoding="utf-8") as x86_config_file :
    for line in x86_config_file :
        line = line.strip()
        if '=' in line : 
            if(len(line) > 0) :
                if(line[0] == '#') :
                    line = line[1:]
                    parts = line.split('=')
                    config_dict[parts[0]][0] = "no use"
                else :
                    parts = line.split('=')
                    config_dict[parts[0]][0] = parts[1]

with open(riscv_config_path, "r", encoding="utf-8") as riscv_config_file :
    for line in riscv_config_file :
        line = line.strip()
        if '=' in line : 
            if(len(line) > 0) :
                if(line[0] == '#') :
                    line = line[1:]
                    parts = line.split('=')
                    config_dict[parts[0]][0] = "no use"
                else :
                    parts = line.split('=')
                    config_dict[parts[0]][1] = parts[1]

wb = openpyxl.Workbook()
ws = wb.create_sheet("config-all")
ws2 = wb.create_sheet("config-noequal")

ws['A1'] = "Config Name"
ws2['A1'] = "Config Name"
ws['B1'] = "x86_64"
ws2['B1'] = "x86_64"
ws['C1'] = "riscv"
ws2['C1'] = "riscv"

for key in config_dict.keys() :
    ws.append([key, config_dict[key][0], config_dict[key][1]])
    if(config_dict[key][0] != config_dict[key][1]) :
        ws2.append([key, config_dict[key][0], config_dict[key][1]])

wb.save("config.xlsx")